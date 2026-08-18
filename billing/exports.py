"""billing/exports.py — Invoice statement exports (Excel + PDF)

Both builders take an already-filtered Invoice queryset (e.g. from the admin
changelist, so it respects whatever month/status/client filter is applied)
and a human-readable period label, and return an HttpResponse file download.
One row per invoice — client name first, no per-product breakdown.
"""
from io import BytesIO
from django.http import HttpResponse


def _product_summary(invoice):
    items = list(invoice.line_items.all())
    if not items:
        return '—'
    first = items[0].description
    extra = len(items) - 1
    if extra > 0:
        return f'{first} (+{extra} more)'
    return first


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------
def build_excel_response(queryset, title_period='All Invoices'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    invoices = list(queryset.prefetch_related('line_items').order_by('invoice_no'))
    total = sum(inv.grand_total for inv in invoices)
    count = len(invoices)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice Statement'

    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'VIRTUAL INSTRUMENTATION & SOFTWARE APPLICATIONS PVT. LTD.'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Invoice Statement — {title_period}'
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center

    ws.merge_cells('A4:C4')
    ws['A4'] = f'Total Invoices: {count}'
    ws['A4'].font = bold

    ws.merge_cells('D4:F4')
    ws['D4'] = f'Total Amount: ₹ {total:,.2f}'
    ws['D4'].font = Font(bold=True, size=12, color='1D4ED8')
    ws['D4'].alignment = Alignment(horizontal='right')

    headers = ['S.No', 'Invoice No', 'Date', 'Client Name', 'Product', 'Amount']
    header_row = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row = header_row + 1
    for i, inv in enumerate(invoices, start=1):
        ws.cell(row=row, column=1, value=i).border = border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=inv.invoice_no).border = border
        ws.cell(row=row, column=3, value=inv.invoice_date.strftime('%d-%m-%Y')).border = border
        ws.cell(row=row, column=4, value=inv.client.name).border = border
        ws.cell(row=row, column=5, value=_product_summary(inv)).border = border
        gt_cell = ws.cell(row=row, column=6, value=float(inv.grand_total))
        gt_cell.number_format = '#,##0.00'
        gt_cell.border = border
        row += 1

    ws.cell(row=row, column=5, value='Grand Total').font = bold
    total_cell = ws.cell(row=row, column=6, value=float(total))
    total_cell.font = bold
    total_cell.number_format = '#,##0.00'

    widths = [6, 12, 12, 26, 42, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{header_row + 1}'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'Invoice_Statement_{title_period.replace(" ", "_")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def build_pdf_response(queryset, title_period='All Invoices'):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    invoices = list(queryset.prefetch_related('line_items').order_by('invoice_no'))
    total = sum(inv.grand_total for inv in invoices)
    count = len(invoices)

    buffer = BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )

    s_company = ParagraphStyle('company', fontName='Helvetica-Bold', fontSize=13, alignment=TA_CENTER)
    s_title   = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=4)
    s_meta    = ParagraphStyle('meta', fontName='Helvetica', fontSize=9)
    s_total   = ParagraphStyle('total', fontName='Helvetica-Bold', fontSize=11, alignment=TA_RIGHT,
                                textColor=colors.HexColor('#1D4ED8'))
    s_cell    = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
    s_cell_r  = ParagraphStyle('cell_r', fontName='Helvetica', fontSize=8, leading=10, alignment=TA_RIGHT)
    s_head    = ParagraphStyle('head', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.white,
                                alignment=TA_CENTER)

    elements = []
    elements.append(Paragraph('VIRTUAL INSTRUMENTATION &amp; SOFTWARE APPLICATIONS PVT. LTD.', s_company))
    elements.append(Paragraph(f'Invoice Statement &mdash; {title_period}', s_title))
    elements.append(Spacer(1, 3 * mm))

    top_row = [[
        Paragraph(f'Total Invoices: <b>{count}</b>', s_meta),
        Paragraph(f'Total Amount: ₹ {total:,.2f}', s_total),
    ]]
    usable_w = page_w - 24 * mm
    top_table = Table(top_row, colWidths=[usable_w * 0.5] * 2)
    top_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -1), 0.75, colors.HexColor('#1D4ED8')),
    ]))
    elements.append(top_table)
    elements.append(Spacer(1, 4 * mm))

    headers = ['S.No', 'Invoice No', 'Date', 'Client Name', 'Product', 'Amount']
    s_cell_c = ParagraphStyle('cell_c', fontName='Helvetica', fontSize=8, leading=10, alignment=TA_CENTER)
    data = [[Paragraph(h, s_head) for h in headers]]
    for i, inv in enumerate(invoices, start=1):
        data.append([
            Paragraph(str(i), s_cell_c),
            Paragraph(inv.invoice_no, s_cell),
            Paragraph(inv.invoice_date.strftime('%d-%m-%Y'), s_cell),
            Paragraph(inv.client.name, s_cell),
            Paragraph(_product_summary(inv), s_cell),
            Paragraph(f'{inv.grand_total:,.2f}', s_cell_r),
        ])
    data.append([
        '', '', '', '',
        Paragraph('<b>Grand Total</b>', s_cell),
        Paragraph(f'<b>{total:,.2f}</b>', s_cell_r),
    ])

    col_widths = [usable_w * w for w in [0.06, 0.12, 0.12, 0.20, 0.36, 0.14]]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('GRID', (0, 0), (-1, -2), 0.4, colors.HexColor('#CCCCCC')),
        ('LINEABOVE', (0, -1), (-1, -1), 0.75, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    filename = f'Invoice_Statement_{title_period.replace(" ", "_")}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response