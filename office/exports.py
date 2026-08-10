"""office/exports.py — Voucher statement exports (Excel + PDF)

Both builders take an already-filtered Voucher queryset (e.g. from the admin
changelist, so it respects whatever month/type/status/search is applied) and
a human-readable period label, and return an HttpResponse file download.
"""
from io import BytesIO
from django.http import HttpResponse


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------
def build_excel_response(queryset, title_period='All Vouchers'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    vouchers = list(queryset)
    total = sum(v.amount for v in vouchers)
    count = len(vouchers)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Voucher Statement'

    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    wrap = Alignment(wrap_text=True, vertical='top')

    # --- Title block ---
    ws.merge_cells('A1:H1')
    ws['A1'] = 'VIRTUAL INSTRUMENTATION & SOFTWARE APPLICATIONS PVT. LTD.'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = center

    ws.merge_cells('A2:H2')
    ws['A2'] = f'Voucher Statement — {title_period}'
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center

    # --- Totals row (TOP, as requested) ---
    ws.merge_cells('A4:D4')
    ws['A4'] = f'Total Vouchers: {count}'
    ws['A4'].font = bold

    ws.merge_cells('E4:H4')
    ws['E4'] = f'Total Amount: ₹ {total:,.2f}'
    ws['E4'].font = Font(bold=True, size=12, color='1D4ED8')
    ws['E4'].alignment = Alignment(horizontal='right')

    # --- Table header ---
    headers = ['Voucher No', 'Type', 'Date', 'Pay To', 'Debit Account', 'Amount', 'Towards', 'Status']
    header_row = 6
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # --- Rows ---
    row = header_row + 1
    for v in vouchers:
        ws.cell(row=row, column=1, value=v.voucher_no).border = border
        ws.cell(row=row, column=2, value=v.get_voucher_type_display()).border = border
        ws.cell(row=row, column=3, value=v.date.strftime('%d-%m-%Y')).border = border
        ws.cell(row=row, column=4, value=v.pay_to).border = border
        ws.cell(row=row, column=5, value=str(v.debit_account)).border = border
        amt_cell = ws.cell(row=row, column=6, value=float(v.amount))
        amt_cell.number_format = '#,##0.00'
        amt_cell.border = border
        towards_cell = ws.cell(row=row, column=7, value=v.towards)
        towards_cell.alignment = wrap
        towards_cell.border = border
        ws.cell(row=row, column=8, value=v.get_status_display()).border = border
        row += 1

    # --- Grand total row at bottom too (bank-statement feel) ---
    ws.cell(row=row, column=5, value='Grand Total').font = bold
    total_cell = ws.cell(row=row, column=6, value=float(total))
    total_cell.font = bold
    total_cell.number_format = '#,##0.00'

    widths = [14, 10, 12, 24, 20, 14, 42, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{header_row + 1}'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'Voucher_Statement_{title_period.replace(" ", "_")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def build_pdf_response(queryset, title_period='All Vouchers'):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    vouchers = list(queryset)
    total = sum(v.amount for v in vouchers)
    count = len(vouchers)

    buffer = BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )

    s_company = ParagraphStyle('company', fontName='Helvetica-Bold', fontSize=13, alignment=TA_CENTER)
    s_title   = ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER, spaceAfter=4)
    s_meta    = ParagraphStyle('meta', fontName='Helvetica', fontSize=9)
    s_total   = ParagraphStyle('total', fontName='Helvetica-Bold', fontSize=11, alignment=TA_RIGHT,
                                textColor=colors.HexColor('#1D4ED8'))
    s_cell    = ParagraphStyle('cell', fontName='Helvetica', fontSize=8, leading=10)
    s_cell_r  = ParagraphStyle('cell_r', fontName='Helvetica', fontSize=8, leading=10, alignment=TA_RIGHT)
    s_head    = ParagraphStyle('head', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.white,
                                alignment=TA_CENTER)

    elements = []
    elements.append(Paragraph('VIRTUAL INSTRUMENTATION &amp; SOFTWARE APPLICATIONS PVT. LTD.', s_company))
    elements.append(Paragraph(f'Voucher Statement &mdash; {title_period}', s_title))
    elements.append(Spacer(1, 3 * mm))

    # --- Totals row (TOP) ---
    top_row = [[
        Paragraph(f'Total Vouchers: <b>{count}</b>', s_meta),
        Paragraph(f'Total Amount: ₹ {total:,.2f}', s_total),
    ]]
    usable_w = page_w - 24 * mm
    top_table = Table(top_row, colWidths=[usable_w * 0.5] * 2)
    top_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -1), 0.75, colors.HexColor('#1D4ED8')),
    ]))
    elements.append(top_table)
    elements.append(Spacer(1, 4 * mm))

    # --- Table ---
    headers = ['Voucher No', 'Type', 'Date', 'Pay To', 'Debit Account', 'Amount', 'Towards', 'Status']
    data = [[Paragraph(h, s_head) for h in headers]]
    for v in vouchers:
        data.append([
            Paragraph(v.voucher_no, s_cell),
            Paragraph(v.get_voucher_type_display(), s_cell),
            Paragraph(v.date.strftime('%d-%m-%Y'), s_cell),
            Paragraph(v.pay_to, s_cell),
            Paragraph(str(v.debit_account), s_cell),
            Paragraph(f'{v.amount:,.2f}', s_cell_r),
            Paragraph(v.towards, s_cell),
            Paragraph(v.get_status_display(), s_cell),
        ])
    data.append([
        '', '', '', '',
        Paragraph('<b>Grand Total</b>', s_cell),
        Paragraph(f'<b>{total:,.2f}</b>', s_cell_r),
        '', '',
    ])

    col_widths = [usable_w * w for w in [0.10, 0.08, 0.09, 0.15, 0.14, 0.10, 0.24, 0.10]]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('GRID', (0, 0), (-1, -2), 0.4, colors.HexColor('#CCCCCC')),
        ('LINEABOVE', (0, -1), (-1, -1), 0.75, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    filename = f'Voucher_Statement_{title_period.replace(" ", "_")}.pdf'
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response